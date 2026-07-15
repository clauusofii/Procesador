import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style.*"
)

st.set_page_config(
    page_title="Demanda Insatisfecha Food Service",
    page_icon="📊",
    layout="wide"
)

LOGO_PATH = "assets/logoAS.png"
MATERIALES_CECINAS_PATH = "Data/materiales_cecinas.xlsx"

MOTIVOS_FUGA = [
    "Contingencia Cliente",
    "Fecha Vencimiento",
    "Fuera de Zona",
    "Local Cerrado",
    "No Despachado",
    "Inconveniente en Ruta",
    "Tiempo de Espera",
    "Pedido Incompleto",
    "Robo o Asalto en Ruta",
    "Termino de Ruta",
    "Problemas con pedido o OC",
    "Problemas técnicos - Sin luz",
]

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #f4f7fb; }
[data-testid="stHeader"] { background: rgba(0,0,0,0); }

.block-container {
    padding-top: 2rem;
    padding-bottom: 2rem;
    max-width: 1250px;
}

.hero {
    background: linear-gradient(135deg, #003B73 0%, #005DAA 100%);
    border-radius: 26px;
    padding: 34px 42px;
    color: white;
    box-shadow: 0 10px 30px rgba(0, 59, 115, 0.22);
    border-bottom: 7px solid #F58220;
    margin-bottom: 32px;
}

.hero-title {
    font-size: 38px;
    font-weight: 900;
    margin-bottom: 8px;
}

.hero-subtitle {
    font-size: 20px;
    font-weight: 650;
    margin-bottom: 10px;
}

.hero-text {
    font-size: 15px;
    color: #e8f2ff;
}

.card {
    background: white;
    border-radius: 22px;
    padding: 26px;
    box-shadow: 0 6px 24px rgba(0,0,0,0.07);
    border: 1px solid #edf1f7;
    margin-bottom: 22px;
}

.section-title {
    font-size: 25px;
    font-weight: 850;
    color: #0B2E4A;
    margin-bottom: 14px;
}

.kpi-card {
    background: white;
    border-radius: 22px;
    padding: 24px 18px;
    min-height: 165px;
    text-align: center;
    box-shadow: 0 6px 24px rgba(0,0,0,0.07);
    border: 1px solid #edf1f7;
}

.kpi-icon {
    font-size: 33px;
    margin-bottom: 8px;
}

.kpi-number {
    font-size: 38px;
    font-weight: 900;
    color: #003B73;
    line-height: 1;
}

.kpi-label {
    font-size: 14px;
    color: #1f2d3d;
    font-weight: 750;
    margin-top: 12px;
}

.kpi-desc {
    font-size: 12px;
    color: #65758a;
    margin-top: 6px;
}

.result-card {
    background: linear-gradient(135deg, #fff7ef 0%, #ffffff 100%);
    border-left: 8px solid #F58220;
    border-radius: 22px;
    padding: 28px;
    box-shadow: 0 6px 24px rgba(0,0,0,0.07);
    margin-bottom: 22px;
}

.result-number {
    font-size: 54px;
    font-weight: 950;
    color: #003B73;
    line-height: 1;
}

.check-list {
    line-height: 2.05;
    font-size: 16px;
}

.footer {
    text-align: center;
    color: #6b7788;
    font-size: 13px;
    margin-top: 36px;
}

.stDownloadButton > button {
    background-color: #003B73;
    color: white;
    border-radius: 14px;
    padding: 0.8rem 1.2rem;
    font-weight: 800;
    border: none;
    width: 100%;
}

.stDownloadButton > button:hover {
    background-color: #005DAA;
    color: white;
    border: none;
}

.small-muted {
    color: #65758a;
    font-size: 13px;
}
</style>
""", unsafe_allow_html=True)


def normalizar_codigo(serie):
    return (
        serie.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )


def crear_resumen_reagendar(df_reagendar):
    """Crea una hoja Resumen tipo tabla dinámica, pero 100% generada con Python."""
    columnas_grupo = [
        "Nombre local",
        "Código Local",
        "Motivo de rechazo",
        "Material SAP",
        "Descripción Item/Material",
    ]

    for columna in columnas_grupo + ["Cantidad faltante (CJ)"]:
        if columna not in df_reagendar.columns:
            return pd.DataFrame(columns=columnas_grupo + ["Cajas a reagendar"])

    resumen = df_reagendar.copy()
    resumen["Cantidad faltante (CJ)"] = pd.to_numeric(
        resumen["Cantidad faltante (CJ)"],
        errors="coerce"
    ).fillna(0)

    resumen = (
        resumen
        .groupby(columnas_grupo, dropna=False, as_index=False)["Cantidad faltante (CJ)"]
        .sum()
        .rename(columns={"Cantidad faltante (CJ)": "Cajas a reagendar"})
        .sort_values(["Nombre local", "Código Local", "Motivo de rechazo", "Material SAP"])
    )

    resumen["Cajas a reagendar"] = resumen["Cajas a reagendar"].round(0).astype(int)
    return resumen


def formatear_hoja(ws, color_header="1F4E79"):
    """Formato general para hojas de datos."""
    header_fill = PatternFill("solid", fgColor=color_header)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 45)


def formatear_resumen(ws):
    """Formato especial para la hoja Resumen."""
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Resumen de pedidos para reagendamiento"
    ws["A1"].font = Font(bold=True, size=16, color="003B73")
    ws["A1"].alignment = Alignment(horizontal="left")

    header_row = 3
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(border_style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            if cell.column in [1, 2, 4]:
                cell.font = Font(bold=True)
            if cell.column == 1:
                cell.fill = light_fill
            if cell.column == 6:
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A4"
    if ws.max_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:F{ws.max_row}"

    widths = {
        "A": 34,
        "B": 16,
        "C": 22,
        "D": 14,
        "E": 42,
        "F": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


left, right = st.columns([1.2, 4.8])

with left:
    st.image(LOGO_PATH, width=230)

with right:
    st.markdown("""
    <div class="hero">
        <div class="hero-title">Procesador de Demanda Insatisfecha</div>
        <div class="hero-subtitle">Fuga involuntaria · Food Service · Sector Cecinas</div>
        <div class="hero-text">
            Identifica pedidos solicitados y despachados, pero no facturados. Es decir, detecta pedidos que no llegaron al cliente luego de haberlos solicitado.
        </div>
    </div>
    """, unsafe_allow_html=True)

st.markdown('<div class="section-title">📂 Carga de archivo</div>', unsafe_allow_html=True)

st.markdown("""
<div class="card">
    <b>Instrucciones:</b><br>
    Sube la data descargada desde Power BI <b>filtrada por el día anterior</b>.
    La aplicación identificará automáticamente los pedidos con fuga involuntaria y generará el listado para su reagendamiento.
</div>
""", unsafe_allow_html=True)

archivo = st.file_uploader(
    "Selecciona el archivo Excel",
    type=["xlsx", "xls"]
)

if archivo is not None:
    try:
        with st.spinner("Procesando reporte..."):
            df_export = pd.read_excel(archivo, sheet_name="Export")

            df_materiales = pd.read_excel(
                MATERIALES_CECINAS_PATH,
                sheet_name="Cecinas"
            )

            materiales_cecinas = set(
                normalizar_codigo(df_materiales["Material"]).dropna()
            )

            df_export["Material SAP normalizado"] = normalizar_codigo(df_export["Material SAP"])
            df_export["Cod Local normalizado"] = normalizar_codigo(df_export["Cod Local"])

            df_ty_gc = df_export[
                df_export["T. Cliente"].isin(["Foodservice"])
            ].copy()

            df_problemas = df_ty_gc[
                df_ty_gc["MotivoRechazo"].isin(MOTIVOS_FUGA)
            ].copy()

            df_problemas["Cantidad recepcionada enteros"] = (
                pd.to_numeric(df_problemas["Cantidad Recepcionada (CJ)"], errors="coerce")
                .fillna(0)
                .round(0)
                .astype(int)
            )

            df_problemas["Cantidad solicitada num"] = (
                pd.to_numeric(df_problemas["Cantidad solicitada (CJ)"], errors="coerce")
                .fillna(0)
            )

            df_problemas["Cantidad faltante"] = (
                df_problemas["Cantidad solicitada num"]
                - df_problemas["Cantidad recepcionada enteros"]
            )

            df_pendientes = df_problemas[
                df_problemas["Cantidad faltante"] > 0
            ].copy()

            df_cecinas_pendientes = df_pendientes[
                df_pendientes["Material SAP normalizado"].isin(materiales_cecinas)
            ].copy()

            cod_locales_afectados = set(
                df_cecinas_pendientes["Cod Local normalizado"].dropna()
            )

            df_pedido_completo_pendiente = df_pendientes[
                df_pendientes["Cod Local normalizado"].isin(cod_locales_afectados)
            ].copy()

            if "Dia entrega" in df_cecinas_pendientes.columns:
                df_cecinas_pendientes["Dia entrega"] = (
                    pd.to_datetime(df_cecinas_pendientes["Dia entrega"], errors="coerce")
                    .dt.strftime("%d-%m-%Y")
                )
                df_cecinas_pendientes.rename(
                    columns={"Dia entrega": "Fecha de incidencia"},
                    inplace=True
                )

            if "Dia entrega" in df_pedido_completo_pendiente.columns:
                df_pedido_completo_pendiente["Dia entrega"] = (
                    pd.to_datetime(df_pedido_completo_pendiente["Dia entrega"], errors="coerce")
                    .dt.strftime("%d-%m-%Y")
                )
                df_pedido_completo_pendiente.rename(
                    columns={"Dia entrega": "Fecha de incidencia"},
                    inplace=True
                )

            columnas_reagendar = [
                "Fecha de incidencia",
                "MotivoRechazo",
                "Sucursal",
                "Cod Local",
                "Nombre local",
                "T. Cliente",
                "Material SAP",
                "Descripción Item/Material",
                "Cantidad faltante",
            ]

            columnas_existentes = [
                col for col in columnas_reagendar
                if col in df_pedido_completo_pendiente.columns
            ]

            df_reagendar = df_pedido_completo_pendiente[columnas_existentes].copy()

            df_reagendar.rename(
                columns={
                    "MotivoRechazo": "Motivo de rechazo",
                    "Cod Local": "Código Local",
                    "Cantidad faltante": "Cantidad faltante (CJ)",
                },
                inplace=True
            )

            if all(col in df_reagendar.columns for col in ["Sucursal", "Nombre local", "Material SAP"]):
                df_reagendar.sort_values(
                    by=["Sucursal", "Nombre local", "Material SAP"],
                    inplace=True
                )

            cajas_cecinas = int(
                round(
                    pd.to_numeric(
                        df_cecinas_pendientes["Cantidad faltante"],
                        errors="coerce"
                    ).fillna(0).sum()
                )
            )

            cajas_reagendar = int(
                round(
                    pd.to_numeric(
                        df_reagendar["Cantidad faltante (CJ)"],
                        errors="coerce"
                    ).fillna(0).sum()
                )
            )

            df_resumen = crear_resumen_reagendar(df_reagendar)

            salida = BytesIO()

            with pd.ExcelWriter(salida, engine="openpyxl") as writer:
                df_resumen.to_excel(writer, sheet_name="Resumen", index=False, startrow=2)
                df_export.to_excel(writer, sheet_name="Data original", index=False)
                df_materiales.to_excel(writer, sheet_name="Materiales cecinas", index=False)
                df_ty_gc.to_excel(writer, sheet_name="Clientes Foodservice", index=False)
                df_problemas.to_excel(writer, sheet_name="Problemas despacho", index=False)
                df_cecinas_pendientes.to_excel(writer, sheet_name="cecinas", index=False)
                df_pedido_completo_pendiente.to_excel(writer, sheet_name="Pedido completo", index=False)
                df_reagendar.to_excel(writer, sheet_name="Re agendar", index=False)

                wb = writer.book
                formatear_resumen(wb["Resumen"])

                for nombre_hoja in wb.sheetnames: 
                    if nombre_hoja != "Resumen":
                        formatear_hoja(wb[nombre_hoja])

            salida.seek(0)

        st.success("✅ Reporte generado correctamente.")

        st.markdown(f"""
        <div class="result-card">
            <div style="font-size: 18px; font-weight: 800; color: #0B2E4A;">
                Resultado del procesamiento
            </div>
            <div style="margin-top: 10px;">
                Clientes con demanda insatisfecha en cecinas:
            </div>
            <div class="result-number">{len(cod_locales_afectados)}</div>
            <div style="font-size: 17px; font-weight: 800; color: #0B2E4A;">
                clientes para reagendar
            </div>
            <div class="small-muted">
                La hoja Re agendar incluye todos los productos pendientes de esos clientes.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="section-title">📊 Resumen del procesamiento</div>', unsafe_allow_html=True)

        col1, col2, col3, col4 = st.columns(4)

        metricas = [
            ("📄", len(df_export), "Data original", "SKU's no facturados de todos los sectores y clientes el día n-1"),
            ("👥", len(df_ty_gc), "Registros Clientes Foodservice", "SKU's no facturados de clientes Food Service"),
            ("📞", cajas_reagendar, "Cajas para Re agendar", "SKU's no entregados en despachos que incluían cecinas"),
            ("🚨", cajas_cecinas, "Cajas de cecinas", "SKU's de cecinas no facturadas"),
        ]

        for col, (icono, numero, label, desc) in zip([col1, col2, col3, col4], metricas):
            with col:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-icon">{icono}</div>
                    <div class="kpi-number">{numero}</div>
                    <div class="kpi-label">{label}</div>
                    <div class="kpi-desc">{desc}</div>
                </div>
                """, unsafe_allow_html=True)

        col_a, col_b = st.columns([1.15, 0.85])

        with col_a:
            st.markdown('<div class="section-title">⚙️ Procesos ejecutados</div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="card check-list">
                ✅ Lectura de data Power BI <br>
                ✅ Lectura automática del maestro de materiales de cecinas<br>
                ✅ Filtrado de Clientes Foodservice<br>
                ✅ Identificación de productos pendientes<br>
                ✅ Detección de clientes con cecinas no entregadas<br>
                ✅ Generación del pedido completo que no se entregó<br>
                ✅ Generación de hoja Re agendar<br>
                ✅ Generación de hoja Resumen agrupada para los pedidos que se deben reagendar
            </div>
            """, unsafe_allow_html=True)

        with col_b:
            st.markdown('<div class="section-title">📥 Descarga</div>', unsafe_allow_html=True)
            fecha = datetime.now().strftime("%Y-%m-%d")

            st.markdown("""
            <div class="card">
                El archivo final incluye:
                <br><br>
                📊 Resumen<br>
                📄 Data original<br>
                📋 Materiales cecinas<br>
                👥 Clientes Foodservices<br>
                📦 Problemas de despacho<br>
                🚨 Pedidos no entregados de cecinas<br>
                🧾 Pedido completo no entregado<br>
                📞 Clientes a quienes Re Agendar
            </div>
            """, unsafe_allow_html=True)

            st.download_button(
                label="⬇ Descargar pedidos a Re Agendar",
                data=salida,
                file_name=f"ReAgendar_{fecha}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("""
        <div class="footer">
            Uso interno Agrosuper · Automatización de reportes · Versión 2.0
        </div>
        """, unsafe_allow_html=True)

    except Exception as e:
        st.error("Ocurrió un error procesando el archivo.")
        st.exception(e)
